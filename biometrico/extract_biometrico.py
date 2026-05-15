"""
Extractor interactivo ZKTeco → raw/biometrico/ y raw/biometrico_parquet/

Genera el mismo prefijo en ambas carpetas: mensual YYYYMMDD (día 01 del mes),
volcado completo YYYYMMDDHHMMSS (solo dígitos).
Menú: mes actual, mes elegido, memoria completa (mismo Excel con inventario de bases del equipo + Metodos_SDK),
memoria y vaciar logs (opción 5), inventario multihoja (opción 6).
También imprime inventario del dispositivo (red, capacidades, firmware, usuarios, etc.).

Uso:
  python biometrico/extract_biometrico.py
  python biometrico/extract_biometrico.py --mes mayo --anio 2026
  python biometrico/extract_biometrico.py --inventario
  python biometrico/extract_biometrico.py --export-inventario
"""

from __future__ import annotations

import argparse
import codecs
import sys
from datetime import datetime, timedelta, timezone
from collections.abc import Callable
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from biometrico.schemas.paths import (
    RAW_BIOMETRICO_DIR,
    RAW_BIOMETRICO_PARQUET_DIR,
    stem_export_memoria_biometrico,
    stem_export_mes_biometrico,
)

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    import polars as pl
    from zk import ZK
except ImportError as e:
    sys.exit(f"[ERROR] Dependencia faltante: {e}\n  Ejecuta: pip install -r requirements.txt")

# Usuarios sin marcar más de este umbral (inventario pantalla + pestañas Usuarios del Excel + filas en Sheet)
DIAS_SIN_MARCAR_INACTIVO = 45
DIAS_ALERTA_AMARILLO_MIN = 2
DIAS_ALERTA_AMARILLO_MAX = 30
MAGENTA_PASTEL = PatternFill(fill_type="solid", fgColor="FFE8B4E8")
AMARILLO_PASTEL = PatternFill(fill_type="solid", fgColor="FFFFF2CC")

MESES_ES: dict[int, str] = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}
MESES_ES_NUM: dict[str, int] = {v: k for k, v in MESES_ES.items()}

DEFAULT_IP = "192.168.100.251"
DEFAULT_PORT = 4370

_METODOS_SDK_INVENTARIO: frozenset[str] = frozenset(
    (
        "get_attendance",
        "get_users",
        "get_templates",
        "get_time",
        "get_network_params",
        "read_sizes",
        "clear_attendance",
        "get_firmware_version",
        "get_serialnumber",
        "get_mac",
        "get_platform",
        "get_device_name",
        "get_fp_version",
        "get_face_version",
        "disable_device",
        "enable_device",
    )
)


def _metodos_sdk_inventario_ordenados() -> list[str]:
    return sorted(_METODOS_SDK_INVENTARIO)


def _safe(fn: Callable[[], object], default: str = "(no disponible)") -> str:
    try:
        v = fn()
        if v is None:
            return default
        return str(v).strip() or default
    except Exception as exc:
        return f"{default} [{type(exc).__name__}: {exc}]"


def _linea_uso_pct(nombre: str, usado: int, cap: int) -> str:
    if cap <= 0:
        return f"  {nombre:14s}: {usado}  (capacidad n/d en SDK)"
    lib = max(0, cap - usado)
    pu, pl = 100.0 * usado / cap, 100.0 * lib / cap
    return (
        f"  {nombre:14s}: {usado:>5}/{cap:<5}  uso {pu:5.1f}%  ·  libre {lib:>5} ({pl:5.1f}%)"
    )


def imprimir_inventario(
    zk: ZK,
    *,
    titulo: str = "Inventario del biométrico (ZK / pyzk)",
    incluir_marcaciones_resumen: bool = True,
) -> None:
    """Lista ‘bases’ útiles del SDK: endpoint, red, memoria, firmware, conteos."""
    sep = "=" * 72
    print(f"\n{sep}\n{titulo}\n{sep}")

    addr = zk.helper.address if hasattr(zk, "helper") else (DEFAULT_IP, DEFAULT_PORT)
    proto = "TCP" if getattr(zk, "tcp", True) else "UDP"
    print("\n[Conexión / endpoint SDK]")
    print(f"  Protocolo     : {proto}")
    print(f"  Host:puerto   : {addr[0]}:{addr[1]}  (comando ZK por defecto 4370)")

    print("\n[Red configurada en el equipo]")
    try:
        net = zk.get_network_params()
        for k, v in net.items():
            print(f"  {k:14s}: {v}")
    except Exception as exc:
        print(f"  (error) {type(exc).__name__}: {exc}")

    print("\n[Memoria y capacidades (read_sizes), con porcentajes de uso]")
    try:
        zk.read_sizes()
    except Exception as exc:
        print(f"  (error read_sizes) {type(exc).__name__}: {exc}")
    print(_linea_uso_pct("Usuarios", zk.users, zk.users_cap))
    print(_linea_uso_pct("Huellas", zk.fingers, zk.fingers_cap))
    print(_linea_uso_pct("Marcaciones", zk.records, zk.rec_cap))
    print(f"  {'Tarjetas':14s}: {zk.cards}")
    print(_linea_uso_pct("Rostros", zk.faces, zk.faces_cap))

    print("\n[Identificación del equipo]")
    print(f"  Firmware      : {_safe(zk.get_firmware_version)}")
    print(f"  Serial        : {_safe(zk.get_serialnumber)}")
    print(f"  Plataforma    : {_safe(zk.get_platform)}")
    print(f"  Nombre        : {_safe(zk.get_device_name)}")
    print(f"  MAC           : {_safe(zk.get_mac)}")
    print(f"  Hora en equipo: {_safe(zk.get_time)}")
    print(f"  Versión FP    : {_safe(zk.get_fp_version)}")
    print(f"  Versión rostro: {_safe(zk.get_face_version)}")

    print("\n[Usuarios en memoria del equipo]")
    users = zk.get_users()
    print(f"  Total usuarios: {len(users)}")
    if users:
        print("  (primeros 8)")
        for u in users[:8]:
            print(f"    · uid={u.uid} user_id={u.user_id!r} name={u.name!r}")
        if len(users) > 8:
            print(f"    … y {len(users) - 8} más")

    if incluir_marcaciones_resumen and users:
        print(
            f"\n[Asistencia: usuarios con más de {DIAS_SIN_MARCAR_INACTIVO} días sin marcar]"
        )
        zk.disable_device()
        try:
            attendances = zk.get_attendance()
        finally:
            zk.enable_device()
        last_map = _ultima_marca_por_user_id(attendances)
        ahora = datetime.now()
        n_inact = sum(
            1
            for u in users
            if _usuario_inactivo_umbral(_ultima_marca_usuario(u, last_map), ahora)
        )
        n_act = len(users) - n_inact
        pct_inact = 100.0 * n_inact / len(users) if users else 0.0
        pct_act = 100.0 * n_act / len(users) if users else 0.0
        print(
            f"  Sin marca reciente (> {DIAS_SIN_MARCAR_INACTIVO} días o nunca): "
            f"{n_inact} de {len(users)}  ({pct_inact:.1f}%)"
        )
        print(f"  Con marca en los últimos {DIAS_SIN_MARCAR_INACTIVO} días: {n_act}  ({pct_act:.1f}%)")
        print(f"  Registros de asistencia en el equipo: {len(attendances)}")

    print("\n[Métodos / ‘bases’ de datos expuestos por pyzk (ZK)]")
    metodos = sorted(
        m
        for m in dir(zk)
        if not m.startswith("_")
        and callable(getattr(zk, m, None))
        and m in _METODOS_SDK_INVENTARIO
    )
    print("  " + ", ".join(metodos))

    print(f"\n{sep}\n")


def _user_map(users: list) -> dict[str, tuple[str, str]]:
    m: dict[str, tuple[str, str]] = {}
    for u in users:
        uid_str = str(u.user_id)
        nombre = (u.name or "").strip() or f"NN-{u.user_id}"
        m[uid_str] = (nombre, uid_str)
    return m


def _build_workbook_marcaciones(
    filas_marcaciones: list[tuple[str, str, datetime]],
    users: list | None,
    attendances_raw: list | None,
    *,
    attendances_para_ultima_marca: list | None = None,
) -> object:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    src_ultima = (
        attendances_para_ultima_marca
        if attendances_para_ultima_marca is not None
        else attendances_raw
    )
    last_map: dict[str, datetime] = _ultima_marca_por_user_id(src_ultima) if src_ultima else {}
    ahora = datetime.now()

    inactivos_por_user_id: set[str] = set()
    if users:
        for u in users:
            ult_u = _ultima_marca_usuario(u, last_map)
            if _usuario_inactivo_umbral(ult_u, ahora):
                inactivos_por_user_id.add(str(u.user_id))

    filas_ord = sorted(
        filas_marcaciones,
        key=lambda r: (str(r[0]), r[2] or datetime.min),
    )

    ws.append(["Número", "Nombre", "Fecha", "Hora"])
    for numero, nombre, ts in filas_ord:
        fecha_c = ts.date() if ts else None
        hora_c = ts.time() if ts else None
        ws.append([numero, nombre, fecha_c, hora_c])
        ri = ws.max_row
        if str(numero) in inactivos_por_user_id:
            for ci in range(1, 5):
                ws.cell(row=ri, column=ci).fill = MAGENTA_PASTEL

    if users:
        wu = wb.create_sheet("Usuarios")
        headers_u = [
            "uid",
            "user_id",
            "nombre",
            "privilegio",
            "grupo",
            "tarjeta",
            "ultima_marca",
            "dias_desde_ultima",
            "estado",
        ]
        wu.append(headers_u)
        for fila in _filas_usuarios_hoja_ordenadas(users, last_map, ahora):
            wu.append(fila)

        _aplicar_rellenos_hoja_usuarios(wu, len(headers_u))

    if attendances_raw is not None:
        um = _user_map(users or [])
        wd = wb.create_sheet("Marcaciones_detalle")
        wd.append(["user_id", "nombre", "tiempo", "status", "punch", "uid"])
        for a in attendances_raw:
            uid_str = str(a.user_id)
            nombre, _ = um.get(uid_str, (f"NN-{a.user_id}", uid_str))
            wd.append([a.user_id, nombre, a.timestamp, a.status, a.punch, a.uid])

    return wb


def _write_excel(
    path: Path,
    filas_marcaciones: list[tuple[str, str, datetime]],
    users: list | None,
    attendances_raw: list | None,
    *,
    attendances_para_ultima_marca: list | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _build_workbook_marcaciones(
        filas_marcaciones,
        users,
        attendances_raw,
        attendances_para_ultima_marca=attendances_para_ultima_marca,
    )
    wb.save(path)


def _write_parquet_mes_uniforme(
    numeros: list[str],
    nombres: list[str],
    tiempos: list[datetime],
    mes: int,
    anio: int,
    pq_path: Path,
) -> None:
    """Misma forma que el respaldo sync_device (MotherDuck / reporte)."""
    pq_path.parent.mkdir(parents=True, exist_ok=True)
    synced_at = datetime.now(timezone.utc).replace(tzinfo=None)
    df = pl.DataFrame(
        {
            "numero": numeros,
            "nombre": nombres,
            "tiempo": tiempos,
            "anio": [anio] * len(tiempos),
            "mes": [mes] * len(tiempos),
            "synced_at": [synced_at] * len(tiempos),
        }
    )
    df.write_parquet(pq_path)


def _write_parquet_memoria_completa(
    numeros: list[str],
    nombres: list[str],
    tiempos: list[datetime],
    pq_path: Path,
) -> None:
    """Un archivo con todas las marcaciones; anio/mes por fila según el timestamp."""
    pq_path.parent.mkdir(parents=True, exist_ok=True)
    synced_at = datetime.now(timezone.utc).replace(tzinfo=None)
    df = pl.DataFrame(
        {
            "numero": numeros,
            "nombre": nombres,
            "tiempo": tiempos,
            "anio": [t.year for t in tiempos],
            "mes": [t.month for t in tiempos],
            "synced_at": [synced_at] * len(tiempos),
        }
    )
    df.write_parquet(pq_path)


def _msg_guardado_ok(stem: str, n: int) -> None:
    """Sin mencionar formatos de archivo."""
    print(
        f"OK ({n} marcaciones) — nombre «{stem}» en "
        f"carpetas «{RAW_BIOMETRICO_DIR.name}» y «{RAW_BIOMETRICO_PARQUET_DIR.name}»."
    )


def _filtrar_por_mes(attendances: list, mes: int, anio: int) -> list:
    return [
        a
        for a in attendances
        if a.timestamp and a.timestamp.year == anio and a.timestamp.month == mes
    ]


def _conectar(ip: str, port: int, password: int) -> ZK:
    zk = ZK(ip, port=port, timeout=15, password=password, force_udp=False, ommit_ping=False)
    zk.connect()
    return zk


def descargar_marcaciones(
    ip: str,
    port: int,
    password: int,
    *,
    mes: int | None,
    anio: int | None,
    todo: bool,
    vaciar_equipo: bool,
    mostrar_inventario: bool = True,
) -> Path | None:
    zk = _conectar(ip, port, password)
    try:
        if mostrar_inventario:
            imprimir_inventario(zk)

        users = zk.get_users()
        um = _user_map(users)

        if not todo and mes is not None and anio is not None:
            zk.disable_device()
            try:
                attendances = zk.get_attendance()
            finally:
                zk.enable_device()
            filtrados = _filtrar_por_mes(attendances, mes, anio)
            if not filtrados:
                print(f"[ERROR] No hay registros para {MESES_ES[mes]}-{anio}.")
                return None
            filtrados.sort(key=lambda a: (str(a.user_id), a.timestamp))
            numeros: list[str] = []
            nombres: list[str] = []
            tiempos: list[datetime] = []
            filas: list[tuple[str, str, datetime]] = []
            for a in filtrados:
                uid_str = str(a.user_id)
                nombre, numero = um.get(uid_str, (f"NN-{a.user_id}", uid_str))
                filas.append((numero, nombre, a.timestamp))
                numeros.append(str(numero))
                nombres.append(nombre)
                tiempos.append(a.timestamp)
            stem = stem_export_mes_biometrico(anio, mes)
            out = RAW_BIOMETRICO_DIR / f"{stem}.xlsx"
            pq = RAW_BIOMETRICO_PARQUET_DIR / f"{stem}.parquet"
            _write_excel(out, filas, users, None, attendances_para_ultima_marca=attendances)
            _write_parquet_mes_uniforme(numeros, nombres, tiempos, mes, anio, pq)
            _msg_guardado_ok(stem, len(filas))
            return out

        # Memoria completa
        zk.disable_device()
        try:
            attendances = zk.get_attendance()
        finally:
            zk.enable_device()

        numeros = []
        nombres = []
        tiempos = []
        filas = []
        for a in sorted(attendances, key=lambda x: (str(x.user_id), x.timestamp)):
            uid_str = str(a.user_id)
            nombre, numero = um.get(uid_str, (f"NN-{a.user_id}", uid_str))
            filas.append((numero, nombre, a.timestamp))
            numeros.append(str(numero))
            nombres.append(nombre)
            tiempos.append(a.timestamp)

        stem = stem_export_memoria_biometrico()
        out = RAW_BIOMETRICO_DIR / f"{stem}.xlsx"
        pq = RAW_BIOMETRICO_PARQUET_DIR / f"{stem}.parquet"

        fingers: list = []
        zk.disable_device()
        try:
            zk.read_sizes()
            try:
                fingers = zk.get_templates()
            except Exception as exc:
                fingers = []
                print(f"[WARN] No se pudieron leer plantillas de huella: {exc}")
        finally:
            zk.enable_device()
        face_rows = _colectar_filas_rostro_plantillas(zk, users)

        out.parent.mkdir(parents=True, exist_ok=True)
        wb = _build_workbook_marcaciones(
            filas, users, attendances, attendances_para_ultima_marca=attendances
        )
        _wb_append_bases_dispositivo_tras_marcaciones(wb, zk, ip, port, users, fingers, face_rows)
        wb.save(out)

        _write_parquet_memoria_completa(numeros, nombres, tiempos, pq)
        _msg_guardado_ok(stem, len(filas))
        print(
            "     Mismo Excel: Sheet, Usuarios, Marcaciones_detalle + Resumen_equipo, Capacidades, Red, "
            "Huellas_plantillas, Rostro_plantillas, Opciones_extendidas, Notas, Metodos_SDK."
        )
        if users:
            print(f"     (Incluye datos de {len(users)} usuarios del equipo.)")

        if vaciar_equipo:
            if filas:
                zk.disable_device()
                try:
                    zk.clear_attendance()
                finally:
                    zk.enable_device()
                zk.read_sizes()
                print(f"Logs de asistencia borrados en el equipo. Marcaciones ahora: {zk.records}")
            else:
                print("(No había marcaciones; no se llama clear_attendance.)")

        return out
    finally:
        zk.disconnect()


def _hex_corto(blob: bytes, max_chars: int = 3200) -> str:
    h = codecs.encode(blob, "hex").decode("ascii")
    if len(h) <= max_chars:
        return h
    return f"{h[:max_chars]}…(total {len(blob)} bytes)"


def _ultima_marca_por_user_id(attendances: list) -> dict[str, datetime]:
    m: dict[str, datetime] = {}
    for a in attendances:
        if not a.timestamp:
            continue
        for key in (str(a.user_id), str(a.uid)):
            t0 = m.get(key)
            if t0 is None or a.timestamp > t0:
                m[key] = a.timestamp
    return m


def _ultima_marca_usuario(u: object, last_map: dict[str, datetime]) -> datetime | None:
    t = last_map.get(str(u.user_id))
    t_uid = last_map.get(str(u.uid))
    if t is None:
        return t_uid
    if t_uid is None:
        return t
    return max(t, t_uid)


def _usuario_inactivo_umbral(ultima: datetime | None, ahora: datetime) -> bool:
    if ultima is None:
        return True
    return (ahora - ultima).total_seconds() > DIAS_SIN_MARCAR_INACTIVO * 86400


def _user_id_sort_key(user_id: object) -> tuple[int, int | str]:
    s = str(user_id).strip()
    try:
        return (0, int(s))
    except ValueError:
        return (1, s.casefold())


def _filas_usuarios_hoja_ordenadas(
    users: list,
    last_map: dict[str, datetime],
    ahora: datetime,
) -> list[list[object]]:
    """Activos arriba (por user_id), inactivos abajo (más días sin marcar primero). uid en columna 1 = 1..n."""
    filas_info: list[tuple[object, datetime | None, int, str, str, str]] = []
    for u in users:
        ult = _ultima_marca_usuario(u, last_map)
        dias_ord = (ahora - ult).days if ult is not None else 1_000_000
        dias_str = "" if ult is None else str((ahora - ult).days)
        estado = "inactivo" if _usuario_inactivo_umbral(ult, ahora) else "activo"
        ult_s = ult.strftime("%Y-%m-%d") if ult else ""
        filas_info.append((u, ult, dias_ord, dias_str, estado, ult_s))

    activos = [t for t in filas_info if t[4] == "activo"]
    inactivos = [t for t in filas_info if t[4] == "inactivo"]
    activos.sort(key=lambda t: _user_id_sort_key(t[0].user_id))
    inactivos.sort(key=lambda t: (-t[2], _user_id_sort_key(t[0].user_id)))

    out: list[list[object]] = []
    seq = 1
    for u, _ult, _d_ord, dias_str, estado, ult_s in activos + inactivos:
        out.append([seq, u.user_id, u.name, u.privilege, u.group_id, u.card, ult_s, dias_str, estado])
        seq += 1
    return out


def _aplicar_rellenos_hoja_usuarios(ws: object, num_cols: int) -> None:
    """Inactivos: magenta. Activos con días sin marcar entre 2 y 30: amarillo pastel."""
    for ri in range(2, ws.max_row + 1):  # type: ignore[attr-defined]
        estado = ws.cell(row=ri, column=9).value  # type: ignore[attr-defined]
        dias_raw = ws.cell(row=ri, column=8).value  # type: ignore[attr-defined]
        fill = None
        if estado == "inactivo":
            fill = MAGENTA_PASTEL
        elif estado == "activo" and dias_raw not in ("", None):
            try:
                d = int(str(dias_raw).strip())
            except ValueError:
                d = -1
            if DIAS_ALERTA_AMARILLO_MIN <= d <= DIAS_ALERTA_AMARILLO_MAX:
                fill = AMARILLO_PASTEL
        if fill is not None:
            for ci in range(1, num_cols + 1):
                ws.cell(row=ri, column=ci).fill = fill  # type: ignore[attr-defined]


def _wb_fill_resumen_equipo(ws: object, zk: ZK, ip: str, port: int) -> None:
    ws.append(["Atributo", "Valor"])  # type: ignore[attr-defined]
    addr = zk.helper.address if hasattr(zk, "helper") else (ip, port)
    rows_kv = [
        ("host", str(addr[0])),
        ("puerto", str(addr[1])),
        ("tcp", str(getattr(zk, "tcp", True))),
        ("firmware", _safe(zk.get_firmware_version)),
        ("serial", _safe(zk.get_serialnumber)),
        ("plataforma", _safe(zk.get_platform)),
        ("nombre_dispositivo", _safe(zk.get_device_name)),
        ("mac", _safe(zk.get_mac)),
        ("hora_equipo", _safe(zk.get_time)),
        ("version_huella_fp", _safe(zk.get_fp_version)),
        ("version_rostro", _safe(zk.get_face_version)),
    ]
    for k, v in rows_kv:
        ws.append([k, v])  # type: ignore[attr-defined]


def _wb_append_capacidades(wb: object, zk: ZK) -> None:
    zk.read_sizes()
    ws_cap = wb.create_sheet("Capacidades")  # type: ignore[attr-defined]
    ws_cap.append(
        [
            "usuarios_actual",
            "usuarios_cap",
            "usuarios_libres",
            "huellas_actual",
            "huellas_cap",
            "marcaciones_actual",
            "marcaciones_cap",
            "tarjetas",
            "rostros_actual",
            "rostros_cap",
        ]
    )
    ws_cap.append(
        [
            zk.users,
            zk.users_cap,
            zk.users_av,
            zk.fingers,
            zk.fingers_cap,
            zk.records,
            zk.rec_cap,
            zk.cards,
            zk.faces,
            zk.faces_cap,
        ]
    )


def _wb_append_red(wb: object, zk: ZK) -> None:
    ws_red = wb.create_sheet("Red")  # type: ignore[attr-defined]
    ws_red.append(["clave", "valor"])
    try:
        for kk, vv in zk.get_network_params().items():
            ws_red.append([kk, vv])
    except Exception as exc:
        ws_red.append(["error", str(exc)])


def _colectar_filas_rostro_plantillas(zk: ZK, users: list) -> list[tuple]:
    face_rows: list[tuple] = []
    if not getattr(zk, "faces", 0):
        return face_rows
    zk.disable_device()
    try:
        for u in users[:100]:
            for tid in (13, 14, 50, 51, 98, 99):
                try:
                    t = zk.get_user_template(u.uid, tid)
                    tpl = getattr(t, "template", None) if t else None
                    if tpl and len(tpl) > 64:
                        face_rows.append(
                            (u.uid, u.user_id, u.name, tid, len(tpl), _hex_corto(tpl, 2400))
                        )
                        break
                except Exception:
                    continue
    finally:
        zk.enable_device()
    return face_rows


def _wb_append_huellas_rostro_opciones_notas(wb: object, zk: ZK, fingers: list, face_rows: list[tuple]) -> None:
    ws_h = wb.create_sheet("Huellas_plantillas")  # type: ignore[attr-defined]
    ws_h.append(["uid_interno", "dedo_id", "valido", "bytes_plantilla", "plantilla_hex_preview"])
    for f in fingers:
        ws_h.append([f.uid, f.fid, f.valid, f.size, _hex_corto(f.template)])

    ws_face = wb.create_sheet("Rostro_plantillas")  # type: ignore[attr-defined]
    ws_face.append(
        ["uid", "user_id", "nombre", "temp_id_probable", "bytes", "plantilla_hex_preview"]
    )
    if face_rows:
        for r in face_rows:
            ws_face.append(list(r))
    else:
        ws_face.append(
            [
                "",
                "",
                "Sin lectura de plantilla facial por SDK (o sin datos).",
                "",
                "",
                "",
            ]
        )

    ws_op = wb.create_sheet("Opciones_extendidas")  # type: ignore[attr-defined]
    ws_op.append(["opcion", "valor"])
    for label, fn in (
        ("ExtendFmt", lambda: zk.get_extend_fmt()),
        ("UserExtFmt", lambda: zk.get_user_extend_fmt()),
        ("FaceFunOn", lambda: zk.get_face_fun_on()),
        ("CompatOldFirmware", lambda: zk.get_compat_old_firmware()),
        ("PinWidth", lambda: zk.get_pin_width()),
    ):
        try:
            ws_op.append([label, fn()])
        except Exception as exc:
            ws_op.append([label, f"(error) {exc}"])

    ws_notas = wb.create_sheet("Notas")  # type: ignore[attr-defined]
    ws_notas.append(["texto"])
    ws_notas.append(
        [
            "Huellas: base FCT_FINGERTMP (get_templates). Rostro: búsqueda heurística con get_user_template "
            "(temp_id típicos 13–14, 50–51, 98–99); depende del firmware."
        ]
    )
    ws_notas.append(
        [
            f"«estado» = inactivo si nunca marcó o la última marca fue hace más de {DIAS_SIN_MARCAR_INACTIVO} días; "
            "esas filas van en magenta pastel en Usuarios. Activos con "
            f"{DIAS_ALERTA_AMARILLO_MIN}–{DIAS_ALERTA_AMARILLO_MAX} días desde la última marca: amarillo pastel. "
            "En Sheet, las marcaciones de user_id inactivos también en magenta."
        ]
    )


def _wb_append_metodos_sdk(wb: object, zk: ZK) -> None:
    ws = wb.create_sheet("Metodos_SDK")  # type: ignore[attr-defined]
    ws.append(["metodo", "callable_en_equipo"])
    for m in _metodos_sdk_inventario_ordenados():
        ok = hasattr(zk, m) and callable(getattr(zk, m, None))
        ws.append([m, "si" if ok else "no"])


def _wb_append_bases_dispositivo_tras_marcaciones(
    wb: object,
    zk: ZK,
    ip: str,
    port: int,
    users: list,
    fingers: list,
    face_rows: list[tuple],
) -> None:
    """Hojas de inventario (mismo criterio que opción 6) al final del libro, sin duplicar Usuarios/Marcaciones."""
    ws_r = wb.create_sheet("Resumen_equipo")  # type: ignore[attr-defined]
    _wb_fill_resumen_equipo(ws_r, zk, ip, port)
    _wb_append_capacidades(wb, zk)
    _wb_append_red(wb, zk)
    _wb_append_huellas_rostro_opciones_notas(wb, zk, fingers, face_rows)
    _wb_append_metodos_sdk(wb, zk)


def exportar_inventario_multihoja(ip: str, port: int, password: int) -> Path:
    """Un libro con varias pestañas: equipo, usuarios, marcaciones, huellas, rostro, opciones, métodos SDK."""
    zk = _conectar(ip, port, password)
    try:
        imprimir_inventario(zk, incluir_marcaciones_resumen=False)
        fingers: list = []
        users: list = []
        attendances: list = []
        zk.disable_device()
        try:
            zk.read_sizes()
            users = zk.get_users()
            attendances = zk.get_attendance()
            try:
                fingers = zk.get_templates()
            except Exception as exc:
                fingers = []
                print(f"[WARN] No se pudieron leer plantillas de huella: {exc}")
        finally:
            zk.enable_device()

        face_rows = _colectar_filas_rostro_plantillas(zk, users)

        last_by_uid = _ultima_marca_por_user_id(attendances)
        ahora_naive = datetime.now()

        stem = f"inventario_equipo_{stem_export_memoria_biometrico()}"
        path = RAW_BIOMETRICO_DIR / f"{stem}.xlsx"
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Resumen_equipo"
        _wb_fill_resumen_equipo(ws0, zk, ip, port)
        _wb_append_capacidades(wb, zk)
        _wb_append_red(wb, zk)

        ws_u = wb.create_sheet("Usuarios")
        headers = [
            "uid",
            "user_id",
            "nombre",
            "privilegio",
            "grupo",
            "tarjeta",
            "ultima_marca",
            "dias_desde_ultima",
            "estado",
        ]
        ws_u.append(headers)
        for fila in _filas_usuarios_hoja_ordenadas(users, last_by_uid, ahora_naive):
            ws_u.append(fila)

        _aplicar_rellenos_hoja_usuarios(ws_u, len(headers))

        ws_m = wb.create_sheet("Marcaciones")
        ws_m.append(["user_id", "nombre", "tiempo", "status", "punch", "uid"])
        um = _user_map(users)
        _sort_ts = datetime(1970, 1, 1)

        def _key_m(a: object) -> tuple:
            ts = a.timestamp if a.timestamp else _sort_ts
            return (str(a.user_id), ts)

        for a in sorted(attendances, key=_key_m):
            n, _ = um.get(str(a.user_id), (f"NN-{a.user_id}", str(a.user_id)))
            ws_m.append([a.user_id, n, a.timestamp, a.status, a.punch, a.uid])

        _wb_append_huellas_rostro_opciones_notas(wb, zk, fingers, face_rows)
        _wb_append_metodos_sdk(wb, zk)

        wb.save(path)
        print(f"Inventario multihoja guardado: {path}")
        return path
    finally:
        zk.disconnect()


def _menu_mes() -> tuple[int, int]:
    hoy = datetime.today()
    print("\nMes (nombre en español, ej. mayo):")
    raw_m = input("> ").strip()
    mes_upper = raw_m.upper()
    if mes_upper not in MESES_ES_NUM:
        raise SystemExit(f"Mes no válido: {raw_m!r}")
    print("Año (Enter = año actual):")
    raw_y = input("> ").strip()
    anio = int(raw_y) if raw_y else hoy.year
    return MESES_ES_NUM[mes_upper], anio


def menu_interactivo(ip: str, port: int, password: int) -> None:
    while True:
        print(
            "\n========================================"
            "\n       Biométrico ZKTeco — Menú"
            "\n========================================"
            "\n"
            "\n  Marcaciones"
            "\n  ----------------------------------------"
            "\n  1) Mes actual"
            "\n  2) Mes específico"
            "\n  3) Memoria completa  (Excel + inventario)"
            "\n"
            "\n  Inventario del equipo"
            "\n  ----------------------------------------"
            "\n  4) Ver en pantalla"
            "\n  5) Exportar a Excel  (multihoja)"
            "\n"
            "\n  0) Salir"
        )
        op = input("\nOpción > ").strip()

        if op == "0":
            print("Chau.")
            return

        hoy = datetime.today()
        if op == "1":
            descargar_marcaciones(ip, port, password, mes=hoy.month, anio=hoy.year, todo=False, vaciar_equipo=False)
        elif op == "2":
            m, y = _menu_mes()
            descargar_marcaciones(ip, port, password, mes=m, anio=y, todo=False, vaciar_equipo=False)
        elif op == "3":
            vaciar = input("¿Vaciar logs del equipo tras guardar? [s/N] ").strip().lower() in ("s", "si", "sí", "y")
            if vaciar:
                print("  Advertencia: se borrarán las marcaciones del reloj (usuarios y huellas NO se afectan).")
                if input("  ¿Confirmar borrado? [s/N] ").strip().lower() not in ("s", "si", "sí", "y"):
                    vaciar = False
            descargar_marcaciones(ip, port, password, mes=None, anio=None, todo=True, vaciar_equipo=vaciar)
        elif op == "4":
            zk = _conectar(ip, port, password)
            try:
                imprimir_inventario(zk)
            finally:
                zk.disconnect()
        elif op == "5":
            exportar_inventario_multihoja(ip, port, password)
        else:
            print("Opción no reconocida.")


def _parse_args() -> argparse.Namespace:
    hoy = datetime.today()
    p = argparse.ArgumentParser(description="ZKTeco → raw/ (menú o CLI)")
    p.add_argument("--ip", default=DEFAULT_IP)
    p.add_argument("--port", default=DEFAULT_PORT, type=int)
    p.add_argument("--password", default=0, type=int)
    p.add_argument("--mes", default=None, help="Mes en español (modo no interactivo)")
    p.add_argument("--anio", default=hoy.year, type=int)
    p.add_argument("--inventario", action="store_true", help="Solo imprimir inventario y salir")
    p.add_argument(
        "--export-inventario",
        action="store_true",
        help="Generar libro multihoja (inventario completo) y salir",
    )
    p.add_argument(
        "--menu",
        action="store_true",
        help="Forzar menú interactivo (útil si stdin no es terminal)",
    )
    p.add_argument(
        "--mes-actual",
        action="store_true",
        help="Descargar el mes en curso sin menú (p. ej. make sync)",
    )
    return p.parse_args()


def main() -> None:
    args = _parse_args()

    if args.inventario:
        zk = _conectar(args.ip, args.port, args.password)
        try:
            imprimir_inventario(zk)
        finally:
            zk.disconnect()
        return

    if args.export_inventario:
        exportar_inventario_multihoja(args.ip, args.port, args.password)
        return

    if args.mes:
        mu = args.mes.upper()
        if mu not in MESES_ES_NUM:
            sys.exit(f"[ERROR] Mes no válido: {args.mes}")
        descargar_marcaciones(
            args.ip,
            args.port,
            args.password,
            mes=MESES_ES_NUM[mu],
            anio=args.anio,
            todo=False,
            vaciar_equipo=False,
        )
        return

    if args.mes_actual:
        hoy = datetime.today()
        descargar_marcaciones(
            args.ip,
            args.port,
            args.password,
            mes=hoy.month,
            anio=hoy.year,
            todo=False,
            vaciar_equipo=False,
        )
        return

    if args.menu or sys.stdin.isatty():
        menu_interactivo(args.ip, args.port, args.password)
        return

    hoy = datetime.today()
    descargar_marcaciones(
        args.ip,
        args.port,
        args.password,
        mes=hoy.month,
        anio=hoy.year,
        todo=False,
        vaciar_equipo=False,
    )


if __name__ == "__main__":
    main()
