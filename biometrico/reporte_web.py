"""
reporte_web.py — Genera reporte HTML interactivo del biométrico.

Uso (desde la carpeta talento/):
    python main.py            → menú principal, opción 1
    python biometrico/reporte_web.py   → ejecución directa
  → genera reports/biometrico/reporte_biometrico.html

Abre el HTML resultante en cualquier navegador.
La configuración de descansos y turnos está en biometrico/horarios.py.
"""

import calendar
import json
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

# Permite ejecutar directamente: python biometrico/reporte_web.py
_ROOT = Path(__file__).parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import polars as pl
from biometrico.horarios import (
    ALMUERZO_MIN,
    DESAYUNO_MIN,
    DESCANSO_MAX_MIN,
    JORNADA_LUNES_VIERNES_MIN,
    JORNADA_SABADO_MIN,
    get_jornada_objetivo_min,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from biometrico.schemas.paths import (
    EXCEL_REPORTS_DIR,
    RAW_BIOMETRICO_DIR,
    RAW_FERIADOS_PATH,
    WEB_REPORT_HTML,
    build_excel_report_path,
)

sys.stdout.reconfigure(encoding="utf-8")

# ── Configuración ──────────────────────────────────────────────────────────────
UMBRAL_DUPLICADO_MIN = 5
RAW_DIR = RAW_BIOMETRICO_DIR
OUTPUT_HTML = WEB_REPORT_HTML

MESES_ES = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}
MESES_ES_INV = {v: k.capitalize() for k, v in MESES_ES.items()}

# ── Helpers de procesamiento ───────────────────────────────────────────────────


def parse_mes(nombre: str) -> tuple[int, int] | None:
    """Extrae (año, mes_num) del nombre del archivo.

    Soporta formatos:
    - 'ENERO-2026.xlsx'       → partes: [MES, AÑO]
    - 'BIO-MARZO-2026.xlsx'   → partes: [PREFIX, MES, AÑO]
    Busca el primer token que sea un mes válido y toma el siguiente como año.
    """
    stem = Path(nombre).stem.upper()
    parts = stem.split("-")
    for i, part in enumerate(parts):
        if part in MESES_ES and i + 1 < len(parts):
            try:
                return int(parts[i + 1]), MESES_ES[part]
            except ValueError:
                continue
    return None


def dedup(filas: list[dict]) -> list[dict]:
    """Elimina marcaciones duplicadas dentro del mismo día (< umbral minutos)."""
    out = []
    for i, f in enumerate(filas):
        if i == 0:
            out.append(f)
            continue
        prev = out[-1]
        if f["fecha"] == prev["fecha"]:
            diff = (f["Tiempo"] - prev["Tiempo"]).total_seconds() / 60
            if diff < UMBRAL_DUPLICADO_MIN:
                continue
        out.append(f)
    return out


def fmt_minutos_legible(minutos: float | None) -> str:
    if minutos is None:
        return "—"
    horas = int(minutos) // 60
    mins = int(round(minutos % 60))
    if mins == 60:
        horas += 1
        mins = 0
    return f"{horas}h {mins:02d}m" if horas > 0 else f"{mins}m"


def _pick_display_name(nombres: list[str]) -> str:
    """Escoge nombre representativo priorizando frecuencia y longitud."""
    limpios = [n for n in nombres if n]
    if not limpios:
        return "SIN NOMBRE"
    counts = Counter(limpios)
    return max(counts.keys(), key=lambda n: (counts[n], len(n), n))


def _to_title_name(nombre: str) -> str:
    """Formatea un nombre para mostrarlo en estilo título."""
    return " ".join(p.capitalize() for p in nombre.split())


def procesar_dia(fecha: date, marcas: list) -> dict:
    """
    Asigna roles a las marcaciones del día y calcula duraciones.
    Misma lógica que main.py.
    """
    n = len(marcas)
    entrada = salida = sal_des = ent_des = sal_alm = ent_alm = None

    if n >= 1:
        entrada = marcas[0]
    if n >= 2:
        salida = marcas[-1]
    if n == 3:
        sal_alm = marcas[1]
    elif n == 4:
        sal_alm, ent_alm = marcas[1], marcas[2]
    elif n == 5:
        sal_des = marcas[1]  # desayuno sin entrada registrada
        sal_alm, ent_alm = marcas[2], marcas[3]
    elif n >= 6:
        sal_des, ent_des = marcas[1], marcas[2]
        sal_alm, ent_alm = marcas[3], marcas[4]

    def dm(a, b):
        return (b - a).total_seconds() / 60 if (a and b) else None

    t_des = dm(sal_des, ent_des)
    t_alm = dm(sal_alm, ent_alm)
    t_jor = dm(entrada, salida)
    t_ef = (t_jor - (t_des or 0) - (t_alm or 0)) if t_jor is not None else None

    def ft(dt):
        return dt.strftime("%H:%M") if dt else None

    def rr(v):
        return round(v, 1) if v is not None else None

    t_descanso_total = (t_des or 0) + (t_alm or 0)
    exceso_desayuno = (
        round(max(0.0, (t_des or 0) - DESAYUNO_MIN), 1) if t_des is not None else None
    )
    exceso_almuerzo = (
        round(max(0.0, (t_alm or 0) - ALMUERZO_MIN), 1) if t_alm is not None else None
    )
    exceso_descanso = (
        round(max(0.0, t_descanso_total - DESCANSO_MAX_MIN), 1)
        if t_jor is not None
        else None
    )
    jornada_objetivo = (
        get_jornada_objetivo_min(fecha.weekday()) if t_ef is not None else None
    )
    horas_extra = (
        round(max(0.0, (t_ef or 0) - jornada_objetivo), 1)
        if t_ef is not None and jornada_objetivo is not None
        else None
    )
    resaltar_descanso = bool(exceso_descanso is not None and exceso_descanso > 5)

    return {
        "entrada": ft(entrada),
        "sal_des": ft(sal_des),
        "ent_des": ft(ent_des),
        "sal_alm": ft(sal_alm),
        "ent_alm": ft(ent_alm),
        "salida": ft(salida),
        "t_desayuno": rr(t_des),
        "t_almuerzo": rr(t_alm),
        "t_jornada": rr(t_jor),
        "t_efectivo": rr(t_ef),
        "jornada_objetivo": jornada_objetivo,
        "exceso_desayuno": exceso_desayuno,
        "exceso_almuerzo": exceso_almuerzo,
        "exceso_descanso": exceso_descanso,
        "horas_extra": horas_extra,
        "resaltar_descanso": resaltar_descanso,
    }


def _normalizar_columnas(df: pl.DataFrame) -> pl.DataFrame:
    """Renombra columnas con encoding corrupto a sus equivalentes correctos.

    Los archivos .xls antiguos a veces entregan 'N�mero' en lugar de 'Número'
    porque el acento no sobrevive la conversión de codepage del formato binario.
    """
    import unicodedata

    mapeo: dict[str, str] = {}
    objetivos = {"Número": r"n.mero", "Nombre": r"nombre", "Tiempo": r"tiempo"}
    for col in df.columns:
        col_norm = unicodedata.normalize("NFKD", col).encode("ascii", "ignore").decode().lower().strip()
        for nombre_correcto, patron in objetivos.items():
            import re
            if re.fullmatch(patron, col_norm) and col != nombre_correcto:
                mapeo[col] = nombre_correcto
    if mapeo:
        df = df.rename(mapeo)
    return df


def _leer_excel(arch: Path) -> pl.DataFrame | None:
    """Lee un archivo .xlsx/.xls detectando automáticamente el nombre de la hoja.

    Intenta primero 'Sheet', luego 'Sheet1', luego la primera hoja disponible.
    Retorna None y loguea el error si el archivo no puede leerse.
    """
    candidatos = ["Sheet", "Sheet1", "Hoja1", "Hoja 1"]
    df = None
    for nombre_hoja in candidatos:
        try:
            df = pl.read_excel(arch, sheet_name=nombre_hoja)
            break
        except Exception:
            continue
    if df is None:
        try:
            df = pl.read_excel(arch)
        except Exception as exc:
            print(f"  [ERROR] No se pudo leer {arch.name}: {exc}")
            return None
    return _normalizar_columnas(df)


def cargar_feriados() -> dict:
    """Carga feriados desde Excel. Retorna {fecha: motivo}."""
    if not RAW_FERIADOS_PATH.exists():
        print(f"  [WARN] Archivo de feriados no encontrado: {RAW_FERIADOS_PATH}")
        return {}
    df = pl.read_excel(RAW_FERIADOS_PATH, sheet_name="feriados")
    return {
        row["Fecha"]: row["Motivo del Feriado"]
        for row in df.to_dicts()
        if row.get("Fecha") and row.get("Motivo del Feriado")
    }


def generar_excel_colaborador(
    colaborador: str,
    mes_key: str,
    mes_label: str,
    mes_data: dict,
) -> str:
    workbook = Workbook()
    resumen_ws = workbook.active
    resumen_ws.title = "Resumen"
    detalle_ws = workbook.create_sheet("Detalle")

    resumen_ws.append(["Reporte biometrico", colaborador])
    resumen_ws.append(["Mes", mes_label])
    resumen_ws.append([])
    resumen = mes_data["resumen"]
    resumen_rows = [
        ("Dias completos", resumen["completos"]),
        ("Dias incompletos", resumen["incompletos"]),
        ("Dias libres o ausentes", resumen["libres"]),
        ("Total horas efectivas", fmt_minutos_legible(resumen["total_efectivo"])),
        ("Total jornada", fmt_minutos_legible(resumen["total_jornada"])),
        ("Promedio efectivo", fmt_minutos_legible(resumen["prom_efectivo"])),
        ("Promedio jornada", fmt_minutos_legible(resumen["prom_jornada"])),
        (
            "Exceso total descanso sobre 45 min",
            fmt_minutos_legible(resumen["total_exceso_descanso"]),
        ),
        ("Horas extra totales", fmt_minutos_legible(resumen["total_horas_extra"])),
        (
            "Jornada objetivo",
            f"L-V {JORNADA_LUNES_VIERNES_MIN // 60}h | Sab {JORNADA_SABADO_MIN // 60}h",
        ),
    ]
    for row in resumen_rows:
        resumen_ws.append(list(row))

    headers = [
        "Fecha",
        "Dia",
        "Estado",
        "Entrada",
        "S. Desayuno",
        "E. Desayuno",
        "S. Almuerzo",
        "E. Almuerzo",
        "Salida",
        "Desayuno",
        "Almuerzo",
        "Exceso 45m",
        "Jornada objetivo",
        "Efectivo",
        "Jornada",
        "Horas extra",
    ]
    detalle_ws.append(headers)

    alert_fill = PatternFill(fill_type="solid", fgColor="FCE7F3")
    alert_font = Font(color="B91C1C", bold=True)

    for day in mes_data["dias"]:
        detalle_ws.append(
            [
                day["fecha"],
                day["dow"],
                day["estado"],
                day["entrada"],
                day["sal_des"],
                day["ent_des"],
                day["sal_alm"],
                day["ent_alm"],
                day["salida"],
                fmt_minutos_legible(day["t_desayuno"]),
                fmt_minutos_legible(day["t_almuerzo"]),
                fmt_minutos_legible(day["exceso_descanso"]),
                fmt_minutos_legible(day["jornada_objetivo"]),
                fmt_minutos_legible(day["t_efectivo"]),
                fmt_minutos_legible(day["t_jornada"]),
                fmt_minutos_legible(day["horas_extra"]),
            ]
        )
        if day.get("resaltar_descanso"):
            row_idx = detalle_ws.max_row
            for cell in detalle_ws[row_idx]:
                cell.fill = alert_fill
                cell.font = alert_font

    for worksheet in (resumen_ws, detalle_ws):
        for row in worksheet.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                current = worksheet.column_dimensions[cell.column_letter].width or 0
                worksheet.column_dimensions[cell.column_letter].width = max(
                    current, len(value) + 2
                )

    output_path = build_excel_report_path(colaborador, mes_key)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path.relative_to(WEB_REPORT_HTML.parent).as_posix()


# ── Carga y procesamiento de todos los archivos ────────────────────────────────


def cargar_datos() -> dict:
    archivos = sorted({*RAW_DIR.glob("*.xlsx"), *RAW_DIR.glob("*.xls")})
    archivos = [a for a in archivos if not a.name.startswith("~$")]
    if not archivos:
        print(f"[ERROR] No hay archivos .xlsx en {RAW_DIR.resolve()}")
        sys.exit(1)

    meses_map = {parse_mes(a.name): a for a in archivos if parse_mes(a.name)}
    if not meses_map:
        print("[ERROR] Ningún archivo tiene el formato MESNAME-YYYY.xlsx")
        sys.exit(1)

    feriados = cargar_feriados()
    datos_por_clave: dict[str, dict[str, dict]] = defaultdict(dict)
    nombre_por_clave: dict[str, str] = {}
    meses_procesados: set[tuple[int, int]] = set()

    for (anio, mes_n), arch in sorted(meses_map.items()):
        mes_key = f"{anio}-{mes_n:02d}"
        print(f"  → Procesando {arch.name}  ({mes_key})")

        try:
            df = _leer_excel(arch)
            if df is None:
                continue

            columnas = set(df.columns)
            for col_req in ("Nombre", "Número", "Tiempo"):
                if col_req not in columnas:
                    raise ValueError(
                        f"columna '{col_req}' no encontrada. "
                        f"Columnas disponibles: {sorted(columnas)}"
                    )

            # Descartar filas con nombres corruptos (bytes de control o garbage del .xls)
            df = df.filter(~pl.col("Nombre").str.contains(r"[\x00-\x1f\xff]"))
            # Normalizar Tiempo: xlsx lo entrega como Datetime, xls como String "DD/MM/YYYY H:MM:SS"
            if df["Tiempo"].dtype == pl.String:
                df = df.with_columns(
                    pl.col("Tiempo").str.strptime(pl.Datetime, "%d/%m/%Y %H:%M:%S")
                )
            df = df.with_columns(
                [
                    pl.col("Nombre")
                    .cast(pl.String)
                    .str.replace_all(r"\s+", " ")
                    .str.strip_chars()
                    .alias("nombre_norm"),
                    pl.col("Número")
                    .cast(pl.String)
                    .str.replace_all(r"\s+", "")
                    .str.strip_chars()
                    .alias("numero_norm"),
                    pl.col("Tiempo").cast(pl.Date).alias("fecha"),
                ]
            ).with_columns(
                pl.when(pl.col("numero_norm").is_not_null() & (pl.col("numero_norm") != ""))
                .then(pl.format("id:{}", pl.col("numero_norm")))
                .otherwise(pl.format("nom:{}", pl.col("nombre_norm")))
                .alias("colab_key")
            )

            claves = sorted(df["colab_key"].unique().to_list())

            for colab_key in claves:
                sub_df = df.filter(pl.col("colab_key") == colab_key)
                nombre_display = _pick_display_name(sub_df["nombre_norm"].to_list())
                nombre_por_clave[colab_key] = nombre_display

                sub = sub_df.sort(["fecha", "Tiempo"]).to_dicts()
                sub = dedup(sub)

                dias_marcas: dict[date, list] = defaultdict(list)
                for f in sub:
                    dias_marcas[f["fecha"]].append(f["Tiempo"])

                _, total_dias = calendar.monthrange(anio, mes_n)
                dias_mes = []

                for d in range(1, total_dias + 1):
                    fecha = date(anio, mes_n, d)
                    dow = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][fecha.weekday()]
                    marcas = sorted(dias_marcas.get(fecha, []))

                    if not marcas:
                        estado = "libre"
                        info = {
                            "entrada": None,
                            "sal_des": None,
                            "ent_des": None,
                            "sal_alm": None,
                            "ent_alm": None,
                            "salida": None,
                            "t_desayuno": None,
                            "t_almuerzo": None,
                            "t_jornada": None,
                            "t_efectivo": None,
                            "jornada_objetivo": None,
                            "exceso_desayuno": None,
                            "exceso_almuerzo": None,
                            "exceso_descanso": None,
                            "horas_extra": None,
                            "resaltar_descanso": False,
                        }
                    else:
                        info = procesar_dia(fecha, marcas)
                        estado = (
                            "completo"
                            if (info["entrada"] and info["salida"])
                            else "incompleto"
                        )

                    dias_mes.append(
                        {
                            "fecha": fecha.strftime("%Y-%m-%d"),
                            "dia": d,
                            "dow": dow,
                            "estado": estado,
                            "motivo_feriado": feriados.get(fecha),
                            **info,
                        }
                    )

                comp = sum(1 for d in dias_mes if d["estado"] == "completo")
                incomp = sum(1 for d in dias_mes if d["estado"] == "incompleto")
                libre = sum(1 for d in dias_mes if d["estado"] == "libre")
                feriados_count = sum(
                    1 for d in dias_mes if d.get("motivo_feriado") is not None
                )
                t_ef = sum(d["t_efectivo"] or 0 for d in dias_mes)
                t_jor = sum(d["t_jornada"] or 0 for d in dias_mes)
                t_exceso_desc = sum(d["exceso_descanso"] or 0 for d in dias_mes)
                t_extra = sum(d["horas_extra"] or 0 for d in dias_mes)

                mes_label = f"{MESES_ES_INV[mes_n]} {anio}"
                resumen = {
                    "completos": comp,
                    "incompletos": incomp,
                    "libres": libre,
                    "feriados": feriados_count,
                    "total_efectivo": round(t_ef),
                    "total_jornada": round(t_jor),
                    "prom_efectivo": round(t_ef / comp) if comp else 0,
                    "prom_jornada": round(t_jor / comp) if comp else 0,
                    "total_exceso_descanso": round(t_exceso_desc),
                    "total_horas_extra": round(t_extra),
                }

                mes_payload = {
                    "dias": dias_mes,
                    "resumen": resumen,
                }
                mes_payload["excel_file"] = generar_excel_colaborador(
                    colaborador=nombre_display,
                    mes_key=mes_key,
                    mes_label=mes_label,
                    mes_data=mes_payload,
                )

                datos_por_clave[colab_key][mes_key] = mes_payload

            meses_procesados.add((anio, mes_n))

        except Exception as exc:
            print(f"  [ERROR] Falló procesando {arch.name}: {exc}")
            print(f"          Este mes será omitido del reporte.")

    meses_lista = [
        {"label": f"{MESES_ES_INV[mes_n]} {anio}", "value": f"{anio}-{mes_n:02d}"}
        for (anio, mes_n) in sorted(meses_procesados)
    ]

    etiquetas_usadas: set[str] = set()
    etiqueta_por_clave: dict[str, str] = {}
    for colab_key in sorted(datos_por_clave.keys()):
        etiqueta_base = nombre_por_clave.get(colab_key, "SIN NOMBRE")
        etiqueta = _to_title_name(etiqueta_base)
        if etiqueta in etiquetas_usadas:
            if colab_key.startswith("id:"):
                etiqueta = f"{etiqueta} ({colab_key[3:]})"
            else:
                etiqueta = f"{etiqueta} ({colab_key})"
        etiquetas_usadas.add(etiqueta)
        etiqueta_por_clave[colab_key] = etiqueta

    datos: dict[str, dict[str, dict]] = {}
    for colab_key, payload_mes in datos_por_clave.items():
        datos[etiqueta_por_clave[colab_key]] = payload_mes

    return {
        "colaboradores": sorted(etiqueta_por_clave.values()),
        "meses": meses_lista,
        "datos": datos,
        "config": {
            "desayuno_min": DESAYUNO_MIN,
            "almuerzo_min": ALMUERZO_MIN,
            "descanso_max_min": DESCANSO_MAX_MIN,
            "jornada_lunes_viernes_min": JORNADA_LUNES_VIERNES_MIN,
            "jornada_sabado_min": JORNADA_SABADO_MIN,
        },
    }


# ── Punto de entrada ───────────────────────────────────────────────────────────


def main() -> None:
    print("Reporte Biométrico Web")
    print("─" * 40)
    EXCEL_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    datos = cargar_datos()

    n_colab = len(datos["colaboradores"])
    n_meses = len(datos["meses"])
    print(f"  Colaboradores: {n_colab}")
    print(
        f"  Meses:         {n_meses}  ({', '.join(m['label'] for m in datos['meses'])})"
    )

    template_path = Path(__file__).parent / "reporte_template.html"
    html_template = template_path.read_text(encoding="utf-8")
    json_str = json.dumps(datos, ensure_ascii=False, separators=(",", ":"))
    html = html_template.replace("__DATA_JSON__", json_str)

    OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_HTML.write_text(html, encoding="utf-8")

    print(f"\n  ✓ Reporte generado: {OUTPUT_HTML.resolve()}")
    print("  Abre el archivo en tu navegador.\n")


if __name__ == "__main__":
    main()
