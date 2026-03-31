"""
processor.py — Lógica de procesamiento del biométrico extraída de reporte_web.py.

El script CLI (reporte_web.py) sigue funcionando sin cambios.
Este módulo provee la misma lógica adaptada para recibir bytes del Excel
(upload HTTP) en lugar de leer desde disco.

Función principal:
    process_excel_file(file_bytes, filename, feriados_path=None) -> ProcessedMonth
"""

from __future__ import annotations

import calendar
import io
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Optional

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ── Imports desde biometrico/ ─────────────────────────────────────────────────
# biometrico/ está al mismo nivel que app/ → subimos un nivel
_BIO_DIR = Path(__file__).resolve().parents[1] / "biometrico"
sys.path.insert(0, str(_BIO_DIR))

from horarios import (  # noqa: E402
    ALMUERZO_MIN,
    DESAYUNO_MIN,
    DESCANSO_MAX_MIN,
    JORNADA_LUNES_VIERNES_MIN,
    JORNADA_SABADO_MIN,
    get_jornada_objetivo_min,
)
from schemas.paths import (  # noqa: E402
    RAW_FERIADOS_PATH,
    slugify_filename,
)

from app.models import ColaboradorResult, DiaData, ProcessedMonth, ResumenData

# ── Constantes ────────────────────────────────────────────────────────────────

UMBRAL_DUPLICADO_MIN = 5

MESES_ES: dict[str, int] = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
MESES_ES_INV: dict[int, str] = {v: k.capitalize() for k, v in MESES_ES.items()}


# ── Helpers (misma lógica que reporte_web.py) ─────────────────────────────────

def parse_mes(nombre: str) -> tuple[int, int] | None:
    """Extrae (año, mes_num) del nombre del archivo.

    Soporta:
    - 'ENERO-2026.xlsx'      → (2026, 1)
    - 'BIO-MARZO-2026.xlsx'  → (2026, 3)
    """
    stem = Path(nombre).stem.upper()
    parts = stem.split("-")
    for i, part in enumerate(parts):
        if part in MESES_ES and i + 1 < len(parts):
            try:
                return int(parts[i + 1]), MESES_ES[part]
            except ValueError:
                continue
    return None


def dedup(filas: list[dict]) -> list[dict]:
    """Elimina marcaciones duplicadas dentro del mismo día (< umbral minutos)."""
    out: list[dict] = []
    for i, f in enumerate(filas):
        if i == 0:
            out.append(f)
            continue
        prev = out[-1]
        if f["fecha"] == prev["fecha"]:
            diff = (f["Tiempo"] - prev["Tiempo"]).total_seconds() / 60
            if diff < UMBRAL_DUPLICADO_MIN:
                continue
        out.append(f)
    return out


def fmt_minutos_legible(minutos: float | None) -> str:
    if minutos is None:
        return "—"
    horas = int(minutos) // 60
    mins = int(round(minutos % 60))
    if mins == 60:
        horas += 1
        mins = 0
    return f"{horas}h {mins:02d}m" if horas > 0 else f"{mins}m"


def _pick_display_name(nombres: list[str]) -> str:
    limpios = [n for n in nombres if n]
    if not limpios:
        return "SIN NOMBRE"
    counts = Counter(limpios)
    return max(counts.keys(), key=lambda n: (counts[n], len(n), n))


def _to_title_name(nombre: str) -> str:
    return " ".join(p.capitalize() for p in nombre.split())


def procesar_dia(fecha: date, marcas: list) -> dict:
    """Asigna roles a las marcaciones del día y calcula duraciones."""
    n = len(marcas)
    entrada = salida = sal_des = ent_des = sal_alm = ent_alm = None

    if n >= 1:
        entrada = marcas[0]
    if n >= 2:
        salida = marcas[-1]
    if n == 3:
        sal_alm = marcas[1]
    elif n == 4:
        sal_alm, ent_alm = marcas[1], marcas[2]
    elif n == 5:
        sal_des = marcas[1]
        sal_alm, ent_alm = marcas[2], marcas[3]
    elif n >= 6:
        sal_des, ent_des = marcas[1], marcas[2]
        sal_alm, ent_alm = marcas[3], marcas[4]

    def dm(a, b):
        return (b - a).total_seconds() / 60 if (a and b) else None

    t_des = dm(sal_des, ent_des)
    t_alm = dm(sal_alm, ent_alm)
    t_jor = dm(entrada, salida)
    t_ef = (t_jor - (t_des or 0) - (t_alm or 0)) if t_jor is not None else None

    def ft(dt):
        return dt.strftime("%H:%M") if dt else None

    def rr(v):
        return round(v, 1) if v is not None else None

    t_descanso_total = (t_des or 0) + (t_alm or 0)
    exceso_desayuno = (
        round(max(0.0, (t_des or 0) - DESAYUNO_MIN), 1) if t_des is not None else None
    )
    exceso_almuerzo = (
        round(max(0.0, (t_alm or 0) - ALMUERZO_MIN), 1) if t_alm is not None else None
    )
    exceso_descanso = (
        round(max(0.0, t_descanso_total - DESCANSO_MAX_MIN), 1)
        if t_jor is not None
        else None
    )
    jornada_objetivo = (
        get_jornada_objetivo_min(fecha.weekday()) if t_ef is not None else None
    )
    horas_extra = (
        round(max(0.0, (t_ef or 0) - jornada_objetivo), 1)
        if t_ef is not None and jornada_objetivo is not None
        else None
    )
    resaltar_descanso = bool(exceso_descanso is not None and exceso_descanso > 5)

    return {
        "entrada": ft(entrada),
        "sal_des": ft(sal_des),
        "ent_des": ft(ent_des),
        "sal_alm": ft(sal_alm),
        "ent_alm": ft(ent_alm),
        "salida": ft(salida),
        "t_desayuno": rr(t_des),
        "t_almuerzo": rr(t_alm),
        "t_jornada": rr(t_jor),
        "t_efectivo": rr(t_ef),
        "jornada_objetivo": jornada_objetivo,
        "exceso_desayuno": exceso_desayuno,
        "exceso_almuerzo": exceso_almuerzo,
        "exceso_descanso": exceso_descanso,
        "horas_extra": horas_extra,
        "resaltar_descanso": resaltar_descanso,
    }


def cargar_feriados(feriados_path: Optional[Path] = None) -> dict[date, str]:
    """Carga feriados desde Excel. Retorna {fecha: motivo}."""
    path = feriados_path or RAW_FERIADOS_PATH
    if not path.exists():
        print(f"  [WARN] Archivo de feriados no encontrado: {path}")
        return {}
    df = pl.read_excel(path, sheet_name="feriados")
    return {
        row["Fecha"]: row["Motivo del Feriado"]
        for row in df.to_dicts()
        if row.get("Fecha") and row.get("Motivo del Feriado")
    }


# ── Generación de Excel por colaborador ───────────────────────────────────────

def generar_excel_colaborador(
    colaborador: str,
    mes_key: str,
    mes_label: str,
    dias: list[DiaData],
    resumen: ResumenData,
    output_dir: Path,
) -> str:
    """Genera el Excel de un colaborador para un mes. Retorna el nombre del archivo."""
    workbook = Workbook()
    resumen_ws = workbook.active
    resumen_ws.title = "Resumen"
    detalle_ws = workbook.create_sheet("Detalle")

    resumen_ws.append(["Reporte biometrico", colaborador])
    resumen_ws.append(["Mes", mes_label])
    resumen_ws.append([])
    resumen_rows = [
        ("Dias completos", resumen["completos"]),
        ("Dias incompletos", resumen["incompletos"]),
        ("Dias libres o ausentes", resumen["libres"]),
        ("Total horas efectivas", fmt_minutos_legible(resumen["total_efectivo"])),
        ("Total jornada", fmt_minutos_legible(resumen["total_jornada"])),
        ("Promedio efectivo", fmt_minutos_legible(resumen["prom_efectivo"])),
        ("Promedio jornada", fmt_minutos_legible(resumen["prom_jornada"])),
        (
            "Exceso total descanso sobre 45 min",
            fmt_minutos_legible(resumen["total_exceso_descanso"]),
        ),
        ("Horas extra totales", fmt_minutos_legible(resumen["total_horas_extra"])),
        (
            "Jornada objetivo",
            f"L-V {JORNADA_LUNES_VIERNES_MIN // 60}h | Sab {JORNADA_SABADO_MIN // 60}h",
        ),
    ]
    for row in resumen_rows:
        resumen_ws.append(list(row))

    headers = [
        "Fecha", "Dia", "Estado",
        "Entrada", "S. Desayuno", "E. Desayuno", "S. Almuerzo", "E. Almuerzo", "Salida",
        "Desayuno", "Almuerzo", "Exceso 45m",
        "Jornada objetivo", "Efectivo", "Jornada", "Horas extra",
    ]
    detalle_ws.append(headers)

    alert_fill = PatternFill(fill_type="solid", fgColor="FCE7F3")
    alert_font = Font(color="B91C1C", bold=True)

    for day in dias:
        detalle_ws.append([
            day["fecha"], day["dia"], day["estado"],
            day["entrada"], day["sal_des"], day["ent_des"],
            day["sal_alm"], day["ent_alm"], day["salida"],
            fmt_minutos_legible(day["t_desayuno"]),
            fmt_minutos_legible(day["t_almuerzo"]),
            fmt_minutos_legible(day["exceso_descanso"]),
            fmt_minutos_legible(day["jornada_objetivo"]),
            fmt_minutos_legible(day["t_efectivo"]),
            fmt_minutos_legible(day["t_jornada"]),
            fmt_minutos_legible(day["horas_extra"]),
        ])
        if day.get("resaltar_descanso"):
            row_idx = detalle_ws.max_row
            for cell in detalle_ws[row_idx]:
                cell.fill = alert_fill
                cell.font = alert_font

    for worksheet in (resumen_ws, detalle_ws):
        for row in worksheet.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                current = worksheet.column_dimensions[cell.column_letter].width or 0
                worksheet.column_dimensions[cell.column_letter].width = max(
                    current, len(value) + 2
                )

    filename = f"reporte_biometrico_{slugify_filename(colaborador)}_{mes_key}.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(output_dir / filename)
    return filename


# ── Función principal ─────────────────────────────────────────────────────────

def process_excel_file(
    file_bytes: bytes,
    filename: str,
    feriados_path: Optional[Path] = None,
) -> ProcessedMonth:
    """Procesa un archivo Excel del biométrico desde bytes (upload HTTP).

    Args:
        file_bytes: Contenido del archivo Excel.
        filename:   Nombre original del archivo (para inferir mes/año).
        feriados_path: Ruta al Excel de feriados. Usa el default si es None.

    Returns:
        ProcessedMonth con todos los colaboradores y sus días procesados.

    Raises:
        ValueError: Si el nombre del archivo no contiene un mes/año válido.
    """
    parsed = parse_mes(filename)
    if parsed is None:
        raise ValueError(
            f"No se pudo inferir mes/año del archivo '{filename}'. "
            "El nombre debe contener un mes en español y un año, p.ej. 'ENERO-2026.xlsx'."
        )
    anio, mes_n = parsed
    mes_key = f"{anio}-{mes_n:02d}"
    mes_label = f"{MESES_ES_INV[mes_n]} {anio}"

    feriados = cargar_feriados(feriados_path)

    # Leer Excel desde bytes (sin escribir a disco)
    df = pl.read_excel(io.BytesIO(file_bytes), sheet_name="Sheet")

    # Descartar filas con nombres corruptos
    df = df.filter(~pl.col("Nombre").str.contains(r"[\x00-\x1f\xff]"))

    # Normalizar Tiempo: xlsx → Datetime, xls → String "DD/MM/YYYY H:MM:SS"
    if df["Tiempo"].dtype == pl.String:
        df = df.with_columns(
            pl.col("Tiempo").str.strptime(pl.Datetime, "%d/%m/%Y %H:%M:%S")
        )

    df = df.with_columns([
        pl.col("Nombre")
            .cast(pl.String)
            .str.replace_all(r"\s+", " ")
            .str.strip_chars()
            .alias("nombre_norm"),
        pl.col("Número")
            .cast(pl.String)
            .str.replace_all(r"\s+", "")
            .str.strip_chars()
            .alias("numero_norm"),
        pl.col("Tiempo").cast(pl.Date).alias("fecha"),
    ]).with_columns(
        pl.when(pl.col("numero_norm").is_not_null() & (pl.col("numero_norm") != ""))
        .then(pl.format("id:{}", pl.col("numero_norm")))
        .otherwise(pl.format("nom:{}", pl.col("nombre_norm")))
        .alias("colab_key")
    )

    claves = sorted(df["colab_key"].unique().to_list())

    # Resolver display_names deduplicados (igual que reporte_web.py)
    nombre_por_clave: dict[str, str] = {}
    for colab_key in claves:
        sub = df.filter(pl.col("colab_key") == colab_key)
        nombre_por_clave[colab_key] = _pick_display_name(sub["nombre_norm"].to_list())

    etiquetas_usadas: set[str] = set()
    etiqueta_por_clave: dict[str, str] = {}
    for colab_key in sorted(claves):
        etiqueta_base = nombre_por_clave.get(colab_key, "SIN NOMBRE")
        etiqueta = _to_title_name(etiqueta_base)
        if etiqueta in etiquetas_usadas:
            suffix = colab_key[3:] if colab_key.startswith("id:") else colab_key
            etiqueta = f"{etiqueta} ({suffix})"
        etiquetas_usadas.add(etiqueta)
        etiqueta_por_clave[colab_key] = etiqueta

    _, total_dias_mes = calendar.monthrange(anio, mes_n)
    result = ProcessedMonth(
        mes_key=mes_key,
        mes_label=mes_label,
        original_file=filename,
    )

    for colab_key in claves:
        display_name = etiqueta_por_clave[colab_key]
        sub_df = df.filter(pl.col("colab_key") == colab_key)
        sub = sub_df.sort(["fecha", "Tiempo"]).to_dicts()
        sub = dedup(sub)

        dias_marcas: dict[date, list] = defaultdict(list)
        for f in sub:
            dias_marcas[f["fecha"]].append(f["Tiempo"])

        dias_mes: list[DiaData] = []
        for d in range(1, total_dias_mes + 1):
            fecha = date(anio, mes_n, d)
            dow = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][fecha.weekday()]
            marcas = sorted(dias_marcas.get(fecha, []))

            if not marcas:
                estado = "libre"
                info: dict = {
                    "entrada": None, "sal_des": None, "ent_des": None,
                    "sal_alm": None, "ent_alm": None, "salida": None,
                    "t_desayuno": None, "t_almuerzo": None, "t_jornada": None,
                    "t_efectivo": None, "jornada_objetivo": None,
                    "exceso_desayuno": None, "exceso_almuerzo": None,
                    "exceso_descanso": None, "horas_extra": None,
                    "resaltar_descanso": False,
                }
            else:
                info = procesar_dia(fecha, marcas)
                estado = (
                    "completo"
                    if (info["entrada"] and info["salida"])
                    else "incompleto"
                )

            dias_mes.append(DiaData(
                fecha=fecha.strftime("%Y-%m-%d"),
                dia=d,
                dow=dow,
                estado=estado,
                motivo_feriado=feriados.get(fecha),
                **info,
            ))

        # Calcular resumen
        comp = sum(1 for d in dias_mes if d["estado"] == "completo")
        incomp = sum(1 for d in dias_mes if d["estado"] == "incompleto")
        libre = sum(1 for d in dias_mes if d["estado"] == "libre")
        feriados_count = sum(1 for d in dias_mes if d.get("motivo_feriado") is not None)
        t_ef = sum(d["t_efectivo"] or 0 for d in dias_mes)
        t_jor = sum(d["t_jornada"] or 0 for d in dias_mes)
        t_exceso = sum(d["exceso_descanso"] or 0 for d in dias_mes)
        t_extra = sum(d["horas_extra"] or 0 for d in dias_mes)

        resumen = ResumenData(
            completos=comp,
            incompletos=incomp,
            libres=libre,
            feriados=feriados_count,
            total_efectivo=round(t_ef),
            total_jornada=round(t_jor),
            prom_efectivo=round(t_ef / comp) if comp else 0,
            prom_jornada=round(t_jor / comp) if comp else 0,
            total_exceso_descanso=round(t_exceso),
            total_horas_extra=round(t_extra),
        )

        result.colaboradores.append(ColaboradorResult(
            colab_key=colab_key,
            display_name=display_name,
            mes_key=mes_key,
            mes_label=mes_label,
            dias=dias_mes,
            resumen=resumen,
        ))

    return result
